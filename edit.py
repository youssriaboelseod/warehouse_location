from openpyxl import load_workbook

wb = load_workbook('demo.xlsx')

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A2'] = 'Tom'
ws['B2'] = 30

ws['A3'] = 'Marry'
ws['B3'] = 29

# Save the file
wb.save("sample.xlsx")